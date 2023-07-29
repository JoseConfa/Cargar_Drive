import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import FiltroArgentina as fa
from openpyxl.utils import get_column_letter
from gspread_formatting import *
import Archivo as ar

gc = gspread.service_account(filename= ar.Credenciales)

sh = gc.create(ar.Nombre)

worksheet = sh.worksheet(ar.Hoja)

sh.share(ar.User, perm_type='user', role='writer')

dataframe = fa.Dataframe_Final
dataframe = dataframe.astype(str)

filas, columnas = dataframe.shape

# Obtener la letra de la columna inicial y final
col_inicial = get_column_letter(1)
col_final = get_column_letter(columnas)

# Construir el rango de hoja de cálculo
range_str= f"{col_inicial}1:{col_final}{filas}"

# Convierte el DataFrame a una lista de listas
data = dataframe.values.tolist()

valor_a_borrar = "nan"
valor_a_reemplazar = ""

# Recorre cada lista dentro de la lista y utiliza la función remove()
for sublist in data:
    for i in range(len(sublist)):
        if sublist[i] == valor_a_borrar:
            sublist[i] = valor_a_reemplazar

# Define el número de columna que deseas alinear al margen derecho (por ejemplo, 3 para la columna D)
num_columna = 3

# Obtiene el rango de celdas para toda la columna
columna_letra = get_column_letter(num_columna + 1)
rango_celdas = f"{columna_letra}:{columna_letra}"

# Define el número de columna que deseas alinear al margen derecho (por ejemplo, 3 para la columna D)
num_columna = 3

# Obtiene el rango de celdas para toda la columna
columna_letra = get_column_letter(num_columna + 1)
rango_celdas = f"{columna_letra}1:{columna_letra}{filas}"

# Establece el formato de alineación a la derecha para la columna
format_request = {
    "repeatCell": {
        "range": {
            "sheetId": worksheet.id,
            "startRowIndex": 0,
            "endRowIndex": filas,
            "startColumnIndex": num_columna,
            "endColumnIndex": num_columna + 1
        },
        "cell": {
            "userEnteredFormat": {
                "horizontalAlignment": "RIGHT"
            }
        },
        "fields": "userEnteredFormat"
    }
}

# Aplica el formato al rango de celdas de la columna
requests = [format_request]
worksheet.spreadsheet.batch_update({"requests": requests})

# Aplica el formato al rango de celdas de la columna
requests = [format_request]
worksheet.spreadsheet.batch_update({"requests": requests})

# Establecer el formato de borde y negrita
cell_format = CellFormat(
    textFormat=TextFormat(bold=True),  # Aplicar formato negrita a los valores
    borders=Borders(
        top=Border(style='SOLID'),     # Agregar borde superior
        bottom=Border(style='SOLID'),  # Agregar borde inferior
        left=Border(style='SOLID'),    # Agregar borde izquierdo
        right=Border(style='SOLID'),   # Agregar borde derecho
    )
)

# Aplicar el formato al rango de celdas
format_cell_range(worksheet, range_str, cell_format)

# Actualizar los datos en la hoja de cálculo
worksheet.update(range_str, data, value_input_option='RAW')